import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import nixio as nix
import neo
from neo.io.nixio import NixIO
from neo.core import (
    Block, ChannelIndex, AnalogSignal, Segment)
from ceed.analysis import CeedDataReader
from quantities import uV, Hz, ms, s
import scipy
import pandas as pd
from openpyxl import load_workbook
from ephys_analysis.spyking_circus_scripts.organize_sc_results import filter_DF, filter_out_DF, \
    get_circus_auto_result_DF, fix_electrode_ids, get_circus_manual_result_DF, get_spikes_by_cluster
from ephys_analysis.spike_train_analysis.unitary_event_scripts import all_binary_patterns, coincidence_plot
from ephys_analysis.ceed_scripts.ceed_stimulus import read_exp_df_from_excel, get_stimulus_signal
import elephant
from elephant.statistics import instantaneous_rate
import elephant.kernels as kernels
import sys
from tqdm import tqdm
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule, DifferentialStyle
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
from ephys_analysis.lfp_processing.filters import custom_round
from mpl_toolkits.mplot3d import Axes3D
from sklearn import decomposition
from sklearn import datasets
from matplotlib.patches import Ellipse, PathPatch
from mpl_toolkits.mplot3d import Axes3D
from spike_lfp_coherence import single_neuron_avg_lineplot_ceed
import mpl_toolkits.mplot3d.art3d as art3d
from itertools import product
from ephys_analysis.spike_train_analysis.unitary_event_scripts import coincidence_plot_odor_trials


def find_rates(date, slice, exp_df, spikes='manual', duration=None,
               patterns=['Odor A', 'Odor B'], header_prefixes=['A_weak', 'B_weak', 'A_strong', 'B_strong'],
               rfs=500*Hz, excel_prefix=None, format_excel=True):

    base_filename = r'D:\Odor quality\{date}\{file}.{ext}'

    excel_file = "rates"
    if excel_prefix is not None:
        excel_file += excel_prefix
    excel = base_filename.format(date=date, file=excel_file, ext="xlsx")
    sheet_name = r'slice' + str(slice)
    book = load_workbook(excel)
    writer = pd.ExcelWriter(excel, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    if spikes == 'manual':
        print("Grabbing manual spikesorting results...")
        slice_folder = date + '___slice' + str(slice) + '_merged'
        gui_folder = slice_folder + r'\\' + slice_folder + 'times.GUI'
        gui_base = r'D:\\Odor quality\\' + date + r'\\' + gui_folder + "\{file}.{ext}"
        # slice_folder = r'SpykingCircus results - SAVED\1-27-20___slice4_merged'
        # gui_folder = slice_folder + r'\1-27-20___slice4_merged' + 'times.GUI'
        # gui_base = r'D:\\' + date + r'\\' + gui_folder + "\{file}.{ext}"
        sc_df = get_circus_manual_result_DF(gui_base, get_electrodes=True, get_groups=False, fs=20000 * Hz)

    exp_df = exp_df[exp_df['pattern'].isin(patterns)]
    exp_df = exp_df[~exp_df['intensity'].isna()]

    print(exp_df)
    intensities = sorted(exp_df['intensity'].unique().tolist())
    num_conditions = len(intensities) * len(patterns)
    num_trials = int(len(exp_df)/num_conditions)

    sts = sc_df["Data"].tolist()
    df = np.empty([len(sts) * num_trials, 2 + 7 * num_conditions])
    df[:] = np.nan

    header1 = ["cluster_ID", "electrode"]
    header2 = ["cluster_ID", "electrode"]
    col_labels = ["baseline_start", "baseline_spikes", "baseline_rate", "odor_start", "odor_spikes", "odor_rate", "norm"]

    for prefix in header_prefixes:
        prefix_header = [prefix] * 7
        header1.extend(prefix_header)
        header2.extend(col_labels)
    header = [np.array(header1), np.array(header2)]

    df = pd.DataFrame(df, columns=header)
    clusters = sc_df["ID"].tolist()
    clusters = [[cluster] * num_trials for cluster in clusters]
    clusters = [item for sublist in clusters for item in sublist]
    electrodes = sc_df["Electrode"].tolist()
    electrodes = [[electrode] * num_trials for electrode in electrodes]
    electrodes = [item for sublist in electrodes for item in sublist]
    df.iloc[:, 0] = clusters
    df.iloc[:, 1] = electrodes

    pbar = tqdm(exp_df.iterrows(), total=exp_df.shape[0], desc="Processing odor trials...", file=sys.stdout)

    trial_counter = 0
    condition_counter = 0

    for index, exp in pbar:
        odor_start, odor_stop = exp['t_start'], exp['t_stop']
        odor_start, odor_stop = np.float(odor_start.strip('s')) * s, np.float(odor_stop.strip('s')) * s
        odor_start, odor_stop = odor_start.rescale(ms), odor_stop.rescale(ms)
        odor_start = custom_round(odor_start, 1000. / rfs) * ms
        odor_stop = custom_round(odor_stop, 1000. / rfs) * ms

        if duration is None:
            duration = (odor_stop - odor_start).rescale(s)

        bl_start = odor_start - duration
        bl_stop = odor_start

        col = 2 + condition_counter * 7
        for i, st in enumerate(sts):
            row = i * num_trials + trial_counter
            st = neo.SpikeTrain(st, t_stop=3600*s, units=ms)  # t_stop is arbitrarily large time, greater than length of actual recording

            df.iloc[row, col+0] = bl_start.rescale(s)
            bl_st = st.time_slice(bl_start, bl_stop)
            bl_spikes = len(bl_st.times)
            if bl_spikes == 0:
                df.iloc[row, col+1:col+3] = 0, 0
            else:
                df.iloc[row, col+1:col+3] = bl_spikes, bl_spikes/np.float(duration)

            df.iloc[row, col+3] = odor_start.rescale(s)
            odor_st = st.time_slice(odor_start, odor_start+duration)
            odor_spikes = len(odor_st.times)
            if odor_spikes == 0:
                df.iloc[row, col+4:col+6] = 0
                continue
            df.iloc[row, col+4:col+6] = odor_spikes, odor_spikes/np.float(duration)
            # if type(int)
            # df.iloc[row, col+6] =
            if bl_spikes != 0:
                df.iloc[row, col+6] = np.round(odor_spikes/bl_spikes,2)
            else:
                df.iloc[row, col+6] = '-'

        condition_counter += 1
        if condition_counter == 4:
            condition_counter = 0
            trial_counter += 1

    pbar.close()
    df = df.fillna('-')
    df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, header=True, index=True)
    sheet = writer.sheets[sheet_name]
    writer.save()

    if format_excel:
        aweak = PatternFill(start_color='d98880', end_color = 'd98880', fill_type = 'solid')  # light red
        bweak = PatternFill(start_color='7fb3da', end_color = '7fb3da', fill_type = 'solid')  # light red
        astrong = PatternFill(start_color='a93226', end_color = 'a93226', fill_type = 'solid')
        bstrong = PatternFill(start_color='2471a3', end_color = '2471a3', fill_type = 'solid')

        aw_rows = sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=4, max_col=10)
        bw_rows = sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=11, max_col=17)
        as_rows = sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=18, max_col=24)
        bs_rows = sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=25, max_col=31)

        for aw_row in aw_rows:
            for cell in aw_row:
                cell.fill = aweak
        for bw_row in bw_rows:
            for cell in bw_row:
                cell.fill = bweak
        for as_row in as_rows:
            for cell in as_row:
                cell.fill = astrong
        for bs_row in bs_rows:
            for cell in bs_row:
                cell.fill = bstrong
        writer.save()


def find_rates_wsr(rates_excel, sheet, conditions=['A_weak', 'B_weak', 'A_strong', 'B_strong'],
                   patterns=['Odor A', 'Odor B'], highlight_sig=True, format_excel=True):

    excel_file = pd.ExcelFile(rates_excel)
    rates_df = excel_file.parse(sheet)
    columns = rates_df.columns.tolist()
    clusters = rates_df.iloc[:, 1].values[1:]
    num_trials = len(rates_df[rates_df["cluster_ID"] == clusters[0]])

    df = np.empty([len(clusters), 1 + 5 * len(conditions)])
    df[:] = np.nan
    header1, header2 = ["cluster_ID"], ["cluster_ID"]
    col_labels = ["odor_start", "baseline_spikes", "odor_spikes", "T", "p"]
    for prefix in conditions:
        prefix_header = [prefix] * 5
        header1.extend(prefix_header)
        header2.extend(col_labels)
    header = [np.array(header1), np.array(header2)]
    wsr_df = pd.DataFrame(df, columns=header)
    wsr_df["cluster_ID"] = clusters
    print(wsr_df)

    for i, cat in enumerate(conditions):
        cat_start_index = columns.index(cat)
        print(rates_df.iloc[:, cat_start_index+1].values[1:])
        bl_spikes = rates_df.iloc[:, cat_start_index+1].values[1:].reshape(int(len(clusters)/num_trials), num_trials)
        odor_spikes = rates_df.iloc[:, cat_start_index+4].values[1:].reshape(int(len(clusters)/num_trials), num_trials)

        Ts, ps = ['-'] * int(len(clusters)), ['-'] * int(len(clusters))
        c = 0
        for x, y in zip(bl_spikes, odor_spikes):
            if np.count_nonzero(x-y) == 0:
                Ts[c*num_trials] = "n/a"
                ps[c*num_trials] = "n/a"
                c += 1
                continue
            T, p = scipy.stats.wilcoxon(y, x, alternative="greater")
            Ts[c*num_trials], ps[c*num_trials] = T, p
            c += 1
        trial_times = rates_df.iloc[:, cat_start_index+3].values[1:]
        wsr_df.iloc[:, 5*i+1] = trial_times
        wsr_df.iloc[:, 5*i+2] = bl_spikes.flatten()
        wsr_df.iloc[:, 5*i+3] = odor_spikes.flatten()
        wsr_df.iloc[:, 5*i+4] = Ts
        wsr_df.iloc[:, 5*i+5] = ps

    writer = pd.ExcelWriter(rates_excel, engine='openpyxl')
    book = load_workbook(rates_excel)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    wsr_sheet = sheet + "_wsr"
    wsr_df.to_excel(writer, sheet_name=wsr_sheet, startrow=0, startcol=0, header=True, index=True)
    wsr_sheet = writer.sheets[wsr_sheet]
    writer.save()

    if format_excel:
        aweak = PatternFill(start_color='d98880', end_color='d98880', fill_type='solid')  # light red
        bweak = PatternFill(start_color='7fb3da', end_color='7fb3da', fill_type='solid')  # light red
        astrong = PatternFill(start_color='a93226', end_color='a93226', fill_type='solid')
        bstrong = PatternFill(start_color='2471a3', end_color='2471a3', fill_type='solid')

        aw_rows = wsr_sheet.iter_rows(min_row=0, max_row=wsr_sheet.max_row, min_col=3, max_col=7)
        bw_rows = wsr_sheet.iter_rows(min_row=0, max_row=wsr_sheet.max_row, min_col=8, max_col=12)
        as_rows = wsr_sheet.iter_rows(min_row=0, max_row=wsr_sheet.max_row, min_col=13, max_col=17)
        bs_rows = wsr_sheet.iter_rows(min_row=0, max_row=wsr_sheet.max_row, min_col=18, max_col=22)

        for aw_row in aw_rows:
            for cell in aw_row:
                cell.fill = aweak
        for bw_row in bw_rows:
            for cell in bw_row:
                cell.fill = bweak
        for as_row in as_rows:
            for cell in as_row:
                cell.fill = astrong
        for bs_row in bs_rows:
            for cell in bs_row:
                cell.fill = bstrong
        writer.save()

    if highlight_sig:
        goldfill = PatternFill(start_color='feeba5', end_color='feeba5', fill_type='solid')
        significant = CellIsRule(operator='lessThan', formula=['.05'], stopIfTrue=True, fill=goldfill)
        wsr_sheet.conditional_formatting.add('G2:G'+str(wsr_sheet.max_row), significant)
        wsr_sheet.conditional_formatting.add('L2:L'+str(wsr_sheet.max_row), significant)
        wsr_sheet.conditional_formatting.add('Q2:Q'+str(wsr_sheet.max_row), significant)
        wsr_sheet.conditional_formatting.add('V2:V'+str(wsr_sheet.max_row), significant)
        writer.save()


def pairwise_odor_cch(exp_df, st1, st2, duration=1*s, odor='Odor A, strong', window=100, bin=5):
    exp_df = exp_df[exp_df['substage'].isin([odor])]
    print(exp_df)

    # fig, ax = plt.subplots(figsize=(12, 8))
    all_odor_st1, all_odor_st2 = [], []
    all_bl_st1, all_bl_st2 = [], []

    y = 1
    for _, experiment in exp_df.iterrows():
        t_start = experiment['t_start']
        t_start = np.float(t_start.strip('s')) * s

        peristim_st1 = st1.time_slice(t_start-duration, t_start+duration)
        peristim_st2 = st2.time_slice(t_start-duration, t_start+duration)

        peristim_st1 = [t - t_start for t in peristim_st1]
        peristim_st2 = [t - t_start for t in peristim_st2]

        # plt.plot(peristim_st1, y*np.ones_like(peristim_st1), 'r.')
        # plt.plot(peristim_st2, y*np.ones_like(peristim_st2), 'b.')
        y += 1

        odor_st1 = st1.time_slice(t_start, t_start+duration)
        odor_st2 = st2.time_slice(t_start, t_start+duration)
        odor_st1 = [t.item() for t in odor_st1]
        odor_st2 = [t.item() for t in odor_st2]
        all_odor_st1 += odor_st1
        all_odor_st2 += odor_st2

        bl_st1 = st1.time_slice(t_start-duration, t_start)
        bl_st2 = st2.time_slice(t_start-duration, t_start)
        bl_st1 = [t.item() for t in bl_st1]
        bl_st2 = [t.item() for t in bl_st2]
        all_bl_st1 += bl_st1
        all_bl_st2 += bl_st2
    # plt.ylabel("Trial #")
    # plt.xlabel("Time from odor onset (ms)")
    # plt.show()

    fig, (ax1, ax2) = plt.subplots(2, 1)
    bin = 5
    low, high = -100, 100
    histo_window = (low, high)
    histo_bins = int((high - low) / bin)

    all_bl_diffs = []
    for spike in all_bl_st1:
        window_start, window_stop = spike - window, spike + window
        in_window = [x for x in all_bl_st2 if x >= window_start and x<= window_stop]
        for spike_time in in_window:
            spike_time = (spike_time - spike)
            all_bl_diffs.append(spike_time)
    counts, bins = np.histogram(all_bl_diffs, bins=histo_bins, range=histo_window)
    ax1.bar(bins[:-1], counts, width=bin, linewidth=1.2, edgecolor='k', facecolor='grey', align='edge')
    # ax1.set_xlabel("time shift (ms)")
    ax1.set_ylabel("# spikes")
    ax1.set_title("Trial-averaged baseline cross-correlation")

    all_odor_diffs = []
    for spike in all_odor_st1:
        window_start, window_stop = spike - window, spike + window
        in_window = [x for x in all_odor_st2 if x >= window_start and x<= window_stop]
        for spike_time in in_window:
            spike_time = (spike_time - spike)
            all_odor_diffs.append(spike_time)
    counts, bins = np.histogram(all_odor_diffs, bins=histo_bins, range=histo_window)
    ax2.bar(bins[:-1], counts, width=bin, linewidth=1.2, edgecolor='k', facecolor='#136fd1', align='edge')
    ax2.set_xlabel("time shift (ms)")
    ax2.set_ylabel("# spikes")
    ax2.set_title("Trial-averaged odor cross-correlation")
    fig = matplotlib.pyplot.gcf()
    fig.set_size_inches(11, 8)
    plt.show()


def pairwise_odor_cch_grid(exp_df, sts, duration=1*s, offset=0*s, odor='Odor A, strong', window=100, bin=5):
    #sts is a list of spiketrains
    exp_df = exp_df[exp_df['substage'].isin([odor])]
    if abs(offset) > abs(duration):
        color = "gray"
    counter = 1
    for i, st1 in enumerate(sts):
        for j, st2 in enumerate(sts):

            all_odor_st1, all_odor_st2 = [], []

            for _, experiment in exp_df.iterrows():
                t_start = experiment['t_start']
                t_start = np.float(t_start.strip('s')) * s

                odor_st1 = st1.time_slice(t_start+offset, t_start+offset+duration)
                odor_st2 = st2.time_slice(t_start+offset, t_start+offset+duration)
                odor_st1 = [t.item() for t in odor_st1]
                odor_st2 = [t.item() for t in odor_st2]
                all_odor_st1 += odor_st1
                all_odor_st2 += odor_st2

            plt.subplot(len(sts), len(sts), counter)
            bin = 5
            low, high = -100, 100
            histo_window = (low, high)
            histo_bins = int((high - low) / bin)

            all_odor_diffs = []
            for spike in all_odor_st1:
                window_start, window_stop = spike - window, spike + window
                in_window = [x for x in all_odor_st2 if x >= window_start and x<= window_stop]
                for spike_time in in_window:
                    spike_time = (spike_time - spike)
                    all_odor_diffs.append(spike_time)
            counts, bins = np.histogram(all_odor_diffs, bins=histo_bins, range=histo_window)
            plt.bar(bins[:-1], counts, width=bin, linewidth=1.2, edgecolor='k', facecolor=color, align='edge')
            plt.title(str(i) + "-" + str(j))
            plt.yticks([])
            plt.xticks([-50, 0, 50])
            plt.xlim([-100, 100])
            plt.grid(True, axis='x')
            counter += 1
    fig = matplotlib.pyplot.gcf()
    fig.set_size_inches(12, 12)
    plt.tight_layout()
    plt.show()


def plot_single_neuron_all_odors(exp_df, st, duration=5*s, bl_duration=1*s, fs=20000*Hz,
                                      odors=['Odor A, weak', 'Odor B, weak', 'Odor A, strong', 'Odor B, strong']):

    exp_df = exp_df[exp_df['substage'].isin(odors)]

    fig, axs = plt.subplots(4, 1)
    colors = ['#d98880', '#7fb3da', '#a93226', '#2471a3']
    for i, odor in enumerate(odors):
        all_odor_st1 = []
        odor_df = exp_df[exp_df['substage'].isin([odor])]
        for _, experiment in odor_df.iterrows():
            t_start = experiment['t_start']
            t_start = np.float(t_start.strip('s')) * s
            peristim_st1 = st.time_slice(t_start-bl_duration, t_start+duration)
            peristim_st1 = [t.rescale(s).item() - t_start.rescale(s).item() for t in peristim_st1]
            all_odor_st1 += peristim_st1

        all_odor_st1 = neo.SpikeTrain(all_odor_st1, units=s, t_start=-1*s, t_stop=duration)
        kernel = kernels.GaussianKernel(sigma=50 * ms)
        ir = instantaneous_rate(all_odor_st1, sampling_period=1 / fs, kernel=kernel)
        ir = ir / len(odor_df)
        # max(ir).item()
        axs[i].plot(ir.times.rescale(s), ir, lw=2, color='k')
        rect = matplotlib.patches.Rectangle((0, 0), 5, 30, linewidth=1, edgecolor=colors[i], facecolor=colors[i], alpha=0.4)
        axs[i].add_patch(rect)
        axs[i].set_ylabel('Firing rate(Hz)', fontsize=10)
        axs[i].set_yticks([0, 10, 20, 30])
        plt.xlabel('Time from odor onset (s)', fontsize=10)
    fig = matplotlib.pyplot.gcf()
    fig.set_size_inches(11, 8)
    plt.show()


def plot_avg_odor_responses(exp_df, sts, duration=5*s, bl_duration=1*s, odor='Odor A, strong', fs=20000*Hz):

    exp_df = exp_df[exp_df['substage'].isin([odor])]
    max_fr, counter = 0, 0
    ir_list = [0] * len(sts)
    fig, ax = plt.subplots(1, 1)
    colors = ['blue', 'orange', 'green', 'red', 'purple', 'brown', 'pink', 'gray', 'olive', 'cyan']
    for st in sts:
        all_odor_st1 = []
        y = 1
        for _, experiment in exp_df.iterrows():
            t_start = experiment['t_start']
            t_start = np.float(t_start.strip('s')) * s
            peristim_st1 = st.time_slice(t_start-bl_duration, t_start+duration)
            peristim_st1 = [t.rescale(s).item() - t_start.rescale(s).item() for t in peristim_st1]
            all_odor_st1 += peristim_st1

            # ax1.scatter(peristim_st1, y*np.ones_like(peristim_st1), color=colors[counter], marker='.')
            y += 1
        all_odor_st1 = neo.SpikeTrain(all_odor_st1, units=s, t_start=-1*s, t_stop=duration)
        kernel = kernels.GaussianKernel(sigma=50 * ms)
        ir = instantaneous_rate(all_odor_st1, sampling_period=1 / fs, kernel=kernel)
        ir = ir / len(exp_df)
        ir_list[counter] = ir

        if max(ir) > max_fr:
            max_fr = max(ir).item()
        counter += 1

    for k, ir in enumerate(ir_list):
        plt.plot(ir.times.rescale(s), ir, lw=2, label='Cluster #' + str(k), color=colors[k])

    rect = matplotlib.patches.Rectangle((0, 0), 5, max_fr,
                                        linewidth=1, edgecolor='b', facecolor='b', alpha=0.2)
    ax.add_patch(rect)
    plt.legend()
    plt.ylabel('Firing rate(Hz)', fontsize=14)
    plt.xlabel('Time from odor onset (s)', fontsize=14)
    fig = matplotlib.pyplot.gcf()
    fig.set_size_inches(11, 8)
    plt.show()


def plot_avg_odor_responses_theta(exp_df, sts, h5_file, odor='Odor A, weak', duration=1*s, bl_duration=1*s, freq=6*Hz, fs=20000*Hz):
    reader = CeedDataReader(h5_file)
    reader.open_h5()
    reader.load_application_data()
    reader.load_mcs_data()

    exp_df = exp_df[exp_df['substage'].isin([odor])]
    theta_duration = (1/freq).rescale(s)
    max_fr, counter = 0, 0
    ir_list, max_args = [0] * len(sts), [0] * len(sts)

    colors = ['blue', 'orange', 'green', 'red', 'purple', 'brown', 'pink', 'gray', 'olive', 'cyan']

    fig = plt.figure()
    ax0 = plt.subplot2grid((5, 1), (0, 0), rowspan=1)
    ax1 = plt.subplot2grid((5, 1), (1, 0), rowspan=4)

    stim = get_stimulus_signal(reader, exp=6, shape='Odor A, circle 1')
    stim = stim.time_slice(stim.t_start, stim.t_start+theta_duration)
    ax0.plot(stim.times, stim, 'b', lw=3)
    ax0.set_xticks([390.619, 390.661, 390.702, 390.744, 390.785])
    ax0.set_xticklabels(['0', '90', '180', '270', '360'])
    for st in sts:
        all_odor_st1 = []
        for _, experiment in exp_df.iterrows():
            t_start = experiment['t_start']
            t_start = np.float(t_start.strip('s')) * s
            for cycle in range(0, int(duration*freq)):
                cycle += 1

                t_stop = t_start + theta_duration
                peristim_st1 = st.time_slice(t_start, t_stop)
                peristim_st1 = [t.rescale(s).item() - t_start.rescale(s).item() for t in peristim_st1]
                all_odor_st1 += peristim_st1
                t_start = t_stop
        all_odor_st1 = neo.SpikeTrain(all_odor_st1, units=s, t_start=0*s, t_stop=theta_duration)
        kernel = kernels.GaussianKernel(sigma=10 * ms)
        ir = instantaneous_rate(all_odor_st1, sampling_period=1 / fs, kernel=kernel)
        ir = ir / (len(exp_df)*6)
        ir_list[counter] = ir
        max_args[counter] = np.argmax(ir)
        if max(ir) > max_fr:
            max_fr = max(ir).item()
        counter += 1

    for k, ir in enumerate(ir_list):
        print(ir.times[max_args[k]].rescale(s))
        ax1.plot(ir.times.rescale(s), ir, lw=2, label='Cluster #' + str(k), color=colors[k])
        ax1.plot([ir.times[max_args[k]].rescale(s), ir.times[max_args[k]].rescale(s)], [0, ir[max_args[k]]], color=colors[k])
    plt.legend()
    plt.ylabel('Firing rate(Hz)', fontsize=14)
    plt.xlabel('Theta cycle-aligned responses', fontsize=14)
    plt.xticks([0, np.float(1/24), np.float(2/24), np.float(3/24), np.float(1/6)], ['0', '90', '180', '270', '360'])
    fig = matplotlib.pyplot.gcf()
    fig.set_size_inches(11, 8)
    plt.show()


def plot_odor_rseponses_theta_raster(exp_df, sts, h5_file, odor='Odor A, weak', duration=1*s, bl_duration=1*s, freq=6*Hz, fs=20000*Hz):

    reader = CeedDataReader(h5_file)
    reader.open_h5()
    reader.load_application_data()
    reader.load_mcs_data()

    exp_df = exp_df[exp_df['substage'].isin([odor])]
    theta_duration = (1/freq).rescale(s)
    ir_list, max_args = [0] * len(sts), [0] * len(sts)

    colors = ['blue', 'orange', 'green', 'red', 'purple', 'brown', 'pink', 'gray', 'olive', 'cyan']

    fig = plt.figure()
    ax0 = plt.subplot2grid((5, 1), (0, 0), rowspan=1)
    ax1 = plt.subplot2grid((5, 1), (1, 0), rowspan=4)

    stim = get_stimulus_signal(reader, exp=6, shape='Odor A, circle 1')
    stim = stim.time_slice(stim.t_start, stim.t_start+theta_duration)
    ax0.plot(stim.times, stim, 'b', lw=3)
    # ax0.set_xticks([390.619, 390.661, 390.702, 390.744, 390.785])
    # ax0.set_xticklabels(['0', '90', '180', '270', '360'])
    counter = 0

    for st in sts:
        all_odor_st1 = []
        exp_counter = 0
        for _, experiment in exp_df.iterrows():
            t_start = experiment['t_start']
            t_start = np.float(t_start.strip('s')) * s
            for cycle in range(0, int(duration*freq)):
                cycle += 1
                y_offset = (cycle-0.5) * np.float(1/int(freq*duration))
                y = exp_counter + y_offset
                t_stop = t_start + theta_duration
                peristim_st1 = st.time_slice(t_start, t_stop)
                peristim_st1 = [t.rescale(s).item() - t_start.rescale(s).item() for t in peristim_st1]
                print(len(peristim_st1))
                ax1.scatter(peristim_st1, y*np.ones_like(peristim_st1), color=colors[counter], marker='.')

                t_start = t_stop
            exp_counter += 1
        counter += 1

    plt.legend()
    plt.ylabel('Trial #', fontsize=14)
    plt.grid(True, axis='y')
    plt.yticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10], ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10'])
    plt.ylim([0, 10])
    plt.xlabel('Theta cycle-aligned responses', fontsize=14)
    plt.xticks([0, np.float(1/24), np.float(2/24), np.float(3/24), np.float(1/6)], ['0', '90', '180', '270', '360'])
    fig = matplotlib.pyplot.gcf()
    fig.set_size_inches(11, 8)
    plt.show()


if __name__ == "__main__":

    # date = '1-27-20'
    date = '1-17-20'
    slice = '4'
    # date = '7-26-20'
    # date = '10-30-20'
    # slice = '2'

    base_filename = r'D:\Odor quality\{date}\{file}.{ext}'

    # base_filename = r'D:\{date}\{file}.{ext}'
    h5_file = date + "___slice" + str(slice) + "_merged"
    h5_file = base_filename.format(date=date, file=h5_file, ext="h5")

    exp_excel = base_filename.format(date=date, file="experiment_df", ext="xlsx")
    exp_sheet = "slice" + str(slice)
    exp_df = read_exp_df_from_excel(exp_excel, exp_sheet)

    slice_folder = date + '___slice' + str(slice) + '_merged'
    gui_folder = slice_folder + r'\\' + slice_folder + 'times.GUI'
    gui_base = r'D:\\' + date + r'\\' + gui_folder + "\{file}.{ext}"
    # result_base = r'D:\\' + date + r'\\' + date + '___slice' + slice + r'_merged\\' +  date + '___slice' + \
    #               slice + r'_mergedtimes.{file}.hdf5'

    # slice_folder = r'SpykingCircus results - SAVED\1-27-20___slice4_merged'
    # gui_folder = slice_folder + r'\1-27-20___slice4_merged' + 'times.GUI'
    # gui_base = r'D:\\' + date + r'\\' + gui_folder + "\{file}.{ext}"

    # sc_df = get_circus_auto_result_DF(result_base, gui_base, fs=20000*Hz)
    # sc_df = get_circus_manual_result_DF(gui_base, get_electrodes=False, get_groups=True, fs=20000 * Hz)
    # complexity_pdf_plot(sc_df, [241*s, 246*s])
    # sc_df = sc_df.iloc[:10, :]

    # find_rates(date, slice, exp_df)
    # find_rates(date, slice, exp_df, excel_prefix="_100ms", duration=100*ms)
    rate_excel = base_filename.format(date=date, file="rates_100ms", ext="xlsx")
    # rate_excel = r'D:\1-27-20\SpykingCircus results - SAVED\rates_saved.xlsx'
    sheet = 'slice' + slice
    find_rates_wsr(rate_excel, sheet)

    # rate_pca(rate_excel, sheet, bad_neurons=bad)

    ### 1/27/20, slice 4
    # st1 = get_spikes_by_cluster(gui_base, 53, 20000*Hz, 1700*s)
    # st2 = get_spikes_by_cluster(gui_base, 76, 20000*Hz, 1700*s)
    # st3 = get_spikes_by_cluster(gui_base, 79, 20000*Hz, 1700*s)
    # st4 = get_spikes_by_cluster(gui_base, 99, 20000*Hz, 1700*s)
    # st5 = get_spikes_by_cluster(gui_base, 114, 20000*Hz, 1700*s)
    # st6 = get_spikes_by_cluster(gui_base, 166, 20000*Hz, 1700*s)
    # st7 = get_spikes_by_cluster(gui_base, 168, 20000*Hz, 1700*s)

    # pairwise_odor_responses(exp_df, st1, st2, odor='Odor A, weak')
    # sts = [st1, st2, st3, st4, st5, st6, st7]

    ### 1/17/20, slice 4
    # st1 = get_spikes_by_cluster(gui_base, 138, 20000*Hz, 1700*s)
    # st2 = get_spikes_by_cluster(gui_base, 31, 20000*Hz, 1700*s)
    # st3 = get_spikes_by_cluster(gui_base, 54, 20000*Hz, 1700*s)
    # st4 = get_spikes_by_cluster(gui_base, 55, 20000*Hz, 1700*s)
    # st5 = get_spikes_by_cluster(gui_base, 75, 20000*Hz, 1700*s)
    # st6 = get_spikes_by_cluster(gui_base, 156, 20000*Hz, 1700*s)
    # st7 = get_spikes_by_cluster(gui_base, 92, 20000*Hz, 1700*s)
    # st8 = get_spikes_by_cluster(gui_base, 93, 20000*Hz, 1700*s)
    # st9 = get_spikes_by_cluster(gui_base, 139, 20000*Hz, 1700*s)
    # st111 = get_spikes_by_cluster(gui_base, 111, 20000*Hz, 1700*s)
    # st91 = get_spikes_by_cluster(gui_base, 91, 20000*Hz, 1700*s)
    # st41 = get_spikes_by_cluster(gui_base, 41, 20000*Hz, 1700*s)
    # st92 = get_spikes_by_cluster(gui_base, 92, 20000*Hz, 1700*s)
    # st95 = get_spikes_by_cluster(gui_base, 95, 20000*Hz, 1700*s)
    # st96 = get_spikes_by_cluster(gui_base, 96, 20000*Hz, 1700*s)
    # st97 = get_spikes_by_cluster(gui_base, 97, 20000*Hz, 1700*s)
    # st98 = get_spikes_by_cluster(gui_base, 98, 20000*Hz, 1700*s)
    # st118 = get_spikes_by_cluster(gui_base, 118, 20000*Hz, 1700*s)
    # e11_gamma_sts = [st98, st118]
    # sts = [st1, st3, st4, st7, st8, st9, st91, st5]
    # sts_aweak = [st7, st41, st92, st6]
    # exp_df = exp_df[~exp_df["experiment"].isin([8, 13])]

    # st1 = elephant.spike_train_surrogates.dither_spikes(st1, 10*ms)[0]
    # st3 = elephant.spike_train_surrogates.dither_spikes(st3, 10*ms)[0]
    # pairwise_odor_cch(exp_df, st1, st3, odor='Odor A, strong')
    # coincidence_plot_odor_trials(h5_file, st1, st3, exp_df, odor="Odor A, strong", binsize=2*ms, winsize=50*ms)

    # pairwise_odor_cch_grid(exp_df, sts, odor='Odor A, weak', duration=500*ms, offset=000*ms)
    # pairwise_odor_cch_grid(exp_df, sts, odor='Odor A, strong', offset=0*s, duration=200*ms)
    # plot_avg_odor_responses(exp_df, sts, odor='Odor B , weak')
    # plot_avg_odor_responses_theta(exp_df, sts, h5_file, odor='Odor A, weak')
    # plot_avg_odor_responses_theta(exp_df, gamma_sts, h5_file, odor='Odor A, weak')
    # exp_df = exp_df[exp_df["experiment"].isin([6, 8, 9, 10])]
    # print(exp_df)
    # plot_odor_rseponses_theta_raster(exp_df, e11_gamma_sts, h5_file, odor='Odor A, weak')

    # exp_df = exp_df[exp_df['experiment']<=10]
    # plot_single_neuron_all_odors(exp_df, st95)
    # print(exp_df)
    # single_neuron_avg_lineplot_ceed(date, slice, 'E11', st95, exp_df, window=5*s, rfs=1000*Hz)
